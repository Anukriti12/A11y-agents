"""
agreement_analysis.py — barrier-labeling reliability + WCAG-vs-barrier divergence

Run AFTER specialists return their filled coding_sheet_specialist_{A,B,C}.xlsx.

Computes:
  1. Label distribution per persona
  2. WCAG-label vs barrier-label divergence table (the key comparison)
  3. Inter-annotator agreement on any cross-labeled validation rows
     (Cohen's kappa, unweighted — labels are nominal)
  4. Severity agreement (quadratic-weighted kappa, Barrier rows only)

Usage:
  python agreement_analysis.py --sheets coding_sheet_specialist_A.xlsx \
      coding_sheet_specialist_B.xlsx coding_sheet_specialist_C.xlsx \
      [--validation validation_labels.xlsx]
"""
import argparse
from collections import defaultdict
from openpyxl import load_workbook

LABELS = ['Barrier', 'No-Barrier', 'Not-Encountered']
LABEL_IDX = {l: i for i, l in enumerate(LABELS)}
SEV = {'Blocking': 3, 'Frustrating': 2, 'Minor': 1, 'N/A': 0, None: 0, '': 0}


def read_sheet(path):
    wb = load_workbook(path, data_only=True)
    ws = wb['Labels']
    # header row is 4, data starts row 5
    rows = []
    for r in range(5, ws.max_row + 1):
        page_id = ws.cell(row=r, column=1).value
        if not page_id:
            continue
        rows.append({
            'page_id': page_id,
            'persona': ws.cell(row=r, column=2).value,
            'wcag': ws.cell(row=r, column=4).value,
            'label': ws.cell(row=r, column=7).value,
            'barrier_desc': ws.cell(row=r, column=8).value,
            'source': ws.cell(row=r, column=9).value,
            'severity': ws.cell(row=r, column=10).value,
            'confidence': ws.cell(row=r, column=11).value,
        })
    return rows


def cohen_kappa(a, b):
    """Unweighted Cohen's kappa for two label lists."""
    assert len(a) == len(b)
    n = len(a)
    if n == 0:
        return None
    cats = sorted(set(a) | set(b))
    idx = {c: i for i, c in enumerate(cats)}
    k = len(cats)
    obs = sum(1 for x, y in zip(a, b) if x == y) / n
    ca = [0]*k; cb = [0]*k
    for x in a: ca[idx[x]] += 1
    for y in b: cb[idx[y]] += 1
    exp = sum((ca[i]/n)*(cb[i]/n) for i in range(k))
    if exp == 1:
        return 1.0
    return (obs - exp) / (1 - exp)


def weighted_kappa(a, b, maxr=3):
    """Quadratic-weighted kappa for ordinal severity 0-3."""
    n = len(a)
    if n == 0:
        return None
    cats = list(range(maxr+1))
    O = [[0]*(maxr+1) for _ in range(maxr+1)]
    for x, y in zip(a, b):
        O[x][y] += 1
    ra = [sum(O[i]) for i in range(maxr+1)]
    rb = [sum(O[i][j] for i in range(maxr+1)) for j in range(maxr+1)]
    num = den = 0
    for i in cats:
        for j in cats:
            w = ((i-j)**2)/(maxr**2)
            e = ra[i]*rb[j]/n
            num += w*O[i][j]
            den += w*e
    if den == 0:
        return 1.0
    return 1 - num/den


def interp(k):
    if k is None: return "n/a"
    if k < 0: return "poor"
    if k < 0.2: return "slight"
    if k < 0.4: return "fair"
    if k < 0.6: return "moderate"
    if k < 0.8: return "substantial"
    return "almost perfect"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sheets', nargs='+', required=True)
    ap.add_argument('--validation', default=None,
                    help="optional second-annotator sheet for the validation sample")
    args = ap.parse_args()

    all_rows = []
    for s in args.sheets:
        all_rows.extend(read_sheet(s))
    print(f"Loaded {len(all_rows)} labeled rows from {len(args.sheets)} sheets\n")

    # 1. Label distribution
    print("="*70)
    print("LABEL DISTRIBUTION PER PERSONA")
    print("="*70)
    dist = defaultdict(lambda: defaultdict(int))
    for r in all_rows:
        dist[r['persona']][r['label']] += 1
    print(f"{'persona':<10}{'Barrier':>9}{'No-Barrier':>12}{'Not-Enc':>10}{'blank':>7}")
    for p in sorted(dist):
        d = dist[p]
        blank = sum(1 for r in all_rows if r['persona']==p and r['label'] not in LABELS)
        print(f"{p:<10}{d.get('Barrier',0):>9}{d.get('No-Barrier',0):>12}{d.get('Not-Encountered',0):>10}{blank:>7}")

    # 2. WCAG vs barrier divergence
    print("\n"+"="*70)
    print("WCAG-LABEL vs BARRIER-LABEL DIVERGENCE  (the key comparison)")
    print("="*70)
    wcag_map = {'passed':'WCAG-pass','failed':'WCAG-fail','inapplicable':'WCAG-N/A'}
    div = defaultdict(int)
    for r in all_rows:
        if r['label'] not in LABELS: continue
        div[(r['wcag'], r['label'])] += 1
    # summarize the interesting divergences
    pass_barrier = sum(n for (w,l),n in div.items() if w=='passed' and l=='Barrier')
    fail_nobar   = sum(n for (w,l),n in div.items() if w=='failed' and l in ('No-Barrier','Not-Encountered'))
    total = sum(div.values())
    print(f"WCAG-pass pages judged a BARRIER for the person:      {pass_barrier}  <- conformance MISSES these")
    print(f"WCAG-fail pages judged No-Barrier/Not-Encountered:    {fail_nobar}  <- conformance OVER-FLAGS these")
    print(f"Total labeled: {total}")
    print(f"\nPer-persona divergence:")
    by_p = defaultdict(lambda: {'pass_barrier':0,'fail_nobar':0,'n':0})
    for r in all_rows:
        if r['label'] not in LABELS: continue
        by_p[r['persona']]['n'] += 1
        if r['wcag']=='passed' and r['label']=='Barrier': by_p[r['persona']]['pass_barrier'] += 1
        if r['wcag']=='failed' and r['label'] in ('No-Barrier','Not-Encountered'): by_p[r['persona']]['fail_nobar'] += 1
    print(f"{'persona':<10}{'pass->barrier':>14}{'fail->nobar':>13}{'n':>5}")
    for p in sorted(by_p):
        d=by_p[p]; print(f"{p:<10}{d['pass_barrier']:>14}{d['fail_nobar']:>13}{d['n']:>5}")

    # 3. Inter-annotator agreement on validation sample
    if args.validation:
        print("\n"+"="*70)
        print("INTER-ANNOTATOR AGREEMENT (validation sample)")
        print("="*70)
        val = {(_['page_id'], _['persona']): _ for _ in read_sheet(args.validation)}
        primary = {(_['page_id'], _['persona']): _ for _ in all_rows}
        common = [k for k in val if k in primary
                  and val[k]['label'] in LABELS and primary[k]['label'] in LABELS]
        if common:
            a = [primary[k]['label'] for k in common]
            b = [val[k]['label'] for k in common]
            agree = sum(1 for x,y in zip(a,b) if x==y)/len(a)
            k = cohen_kappa(a, b)
            print(f"n={len(common)}  raw agreement={agree:.1%}  Cohen κ={k:.3f} ({interp(k)})")
            # severity agreement on Barrier rows both agreed
            sev_a, sev_b = [], []
            for key in common:
                if primary[key]['label']=='Barrier' and val[key]['label']=='Barrier':
                    sev_a.append(SEV.get(primary[key]['severity'],0))
                    sev_b.append(SEV.get(val[key]['severity'],0))
            if len(sev_a) >= 2:
                wk = weighted_kappa(sev_a, sev_b)
                print(f"Severity (Barrier rows, n={len(sev_a)}): weighted κ={wk:.3f} ({interp(wk)})")
            # confusion
            print("\nConfusion (primary rows vs validation):")
            conf = defaultdict(int)
            for x,y in zip(a,b): conf[(x,y)] += 1
            for x in LABELS:
                row = "  ".join(f"{conf.get((x,y),0):>3}" for y in LABELS)
                print(f"  {x:<16} {row}")
            print(f"  {'':16} " + "  ".join(f"{l[:3]:>3}" for l in LABELS))
        else:
            print("No overlapping validation rows found.")
    else:
        print("\n(no --validation sheet supplied; skipping inter-annotator kappa)")

    # 4. Low-confidence + missing-detail flags
    print("\n"+"="*70)
    print("QUALITY FLAGS")
    print("="*70)
    low = [r for r in all_rows if r.get('confidence')=='Low']
    print(f"Low-confidence rows (discuss at consensus): {len(low)}")
    for r in low:
        print(f"  {r['page_id']} {r['persona']}: {r['label']}")
    missing = [r for r in all_rows if r['label']=='Barrier' and not r.get('barrier_desc')]
    if missing:
        print(f"\nBarrier rows missing description ({len(missing)}):")
        for r in missing:
            print(f"  {r['page_id']} {r['persona']}")


if __name__ == '__main__':
    main()
